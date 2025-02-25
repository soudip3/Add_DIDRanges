import authentication
import createDIDRange
import getCallRouting
import assignDID
import openpyxl

CLIENT_ID = "<client_id>"
CLIENT_SECRET = "<client_secret>"
ENVIRONMENT = "<env>" # eg. mypurecloud.com
callRoutingId = "<callroutingId>"
StartDID = ""
EndDID = ""

#get token
access_token = authentication.authentication(CLIENT_ID, CLIENT_SECRET, ENVIRONMENT)

#read xlsx
path = "DID.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
row = sheet_obj.max_row
column = sheet_obj.max_column

# create DID range
for i in range(1, row + 1):
    cell_obj = sheet_obj.cell(row=i, column=1)
    cell_obj_next = sheet_obj.cell(row=i+1, column=1)
    try:
        if int(cell_obj.value)+1 == int(cell_obj_next.value):
            if StartDID == "": 
                StartDID = "+44" + str(cell_obj.value)
            continue
        else:
            if StartDID == "":
                StartDID = "+44" + str(cell_obj.value)
            EndDID = "+44" + str(cell_obj.value)
            createDIDRange.addRanges(access_token, ENVIRONMENT, StartDID, EndDID)
            StartDID = ""
            EndDID = ""
    except:
        if StartDID == "":
            StartDID = "+44" + str(cell_obj.value)
        EndDID = "+44" + str(cell_obj.value)
        createDIDRange.addRanges(access_token, ENVIRONMENT, StartDID, EndDID)
        StartDID = ""
        EndDID = ""

#assign DID to call routing
for j in range(5,int((row/100)+1)):
    responseCallRouting = getCallRouting.callRouting(access_token, callRoutingId, ENVIRONMENT)
    if 100*(j+1)>row:
        for i in range((100*j)+1, row+1):
            cell_obj = sheet_obj.cell(row=i, column=1)
            responseCallRouting["dnis"].append("+44" + str(cell_obj.value))
    else:
        for i in range((100*j)+1, 100*(j+1)+1):
            cell_obj = sheet_obj.cell(row=i, column=1)
            responseCallRouting["dnis"].append("+44" + str(cell_obj.value))
    assignDID.addDIDs(access_token, responseCallRouting, ENVIRONMENT, callRoutingId)




